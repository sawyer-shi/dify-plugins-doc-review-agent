from collections.abc import Generator
from typing import Any
import csv
import json

from openpyxl import load_workbook

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage

from tools.utils import save_upload_to_temp, clean_paths, dual_messages


class RuleLoaderTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        rules_file = tool_parameters.get("rules_file")

        if not rules_file:
            for m in dual_messages(self, "Error: rules_file is required.", {"error": "rules_file is required"}):
                yield m
            return

        required_fields = ["rule_code", "rule_name", "rule_level", "rule_prompt"]
        temp_path = None

        try:
            temp_path, _, ext = save_upload_to_temp(rules_file)
            rows: list[dict[str, str]] = []

            if ext == ".csv":
                loaded = False
                for enc in ["utf-8-sig", "utf-8", "gbk"]:
                    try:
                        with open(temp_path, "r", encoding=enc, newline="") as f:
                            reader = csv.DictReader(f)
                            if not reader.fieldnames:
                                continue
                            missing = [k for k in required_fields if k not in reader.fieldnames]
                            if missing:
                                msg = f"Error: Missing fields in rules csv: {', '.join(missing)}"
                                for m in dual_messages(self, msg, {"error": msg}):
                                    yield m
                                return
                            for row in reader:
                                rows.append({k: str(row.get(k, "")).strip() for k in required_fields})
                        loaded = True
                        break
                    except Exception:
                        continue
                if not loaded:
                    for m in dual_messages(self, "Error: Failed to read rules csv file.", {"error": "Failed to read rules csv file"}):
                        yield m
                    return
            elif ext == ".xlsx":
                wb = load_workbook(temp_path, data_only=True)
                ws = wb.worksheets[0]
                header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                headers = [str(c).strip() if c is not None else "" for c in header_cells]
                header_map = {h: idx for idx, h in enumerate(headers)}
                missing = [k for k in required_fields if k not in header_map]
                if missing:
                    msg = f"Error: Missing fields in rules xlsx: {', '.join(missing)}"
                    for m in dual_messages(self, msg, {"error": msg}):
                        yield m
                    return

                for values in ws.iter_rows(min_row=2, values_only=True):
                    item: dict[str, str] = {}
                    is_empty = True
                    for k in required_fields:
                        val = values[header_map[k]] if header_map[k] < len(values) else ""
                        sval = "" if val is None else str(val).strip()
                        if sval:
                            is_empty = False
                        item[k] = sval
                    if not is_empty:
                        rows.append(item)
            else:
                for m in dual_messages(self, "Error: rules_file must be .csv or .xlsx", {"error": "rules_file must be .csv or .xlsx"}):
                    yield m
                return

            rows = [r for r in rows if r.get("rule_code") and r.get("rule_prompt")]
            if not rows:
                for m in dual_messages(self, "Error: No valid rules loaded from rules_file.", {"error": "No valid rules loaded from rules_file"}):
                    yield m
                return

            checklist = []
            for idx, r in enumerate(rows, start=1):
                checklist.append(
                    f"{idx}. [{r['rule_code']}][{r['rule_level']}] {r['rule_name']}: {r['rule_prompt']}"
                )

            payload = {
                "rules": rows,
                "rules_text": "\n".join(checklist),
            }
            out_text = json.dumps(payload, ensure_ascii=False)
            for m in dual_messages(self, out_text, payload):
                yield m

        finally:
            clean_paths([temp_path] if temp_path else [])
